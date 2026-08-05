#!/usr/bin/env python3
"""
One-off export: reads "Career Library_Updated_0508.xlsx" and writes clean JSON per
tab into prisma/seed-data/career-library/, for prisma/seed.ts to load.

Ignores the "Post-12_Entrance_Exams__India__" tab per instruction (out of scope).
Not part of the app's runtime — rerun manually if the source workbook changes.
"""
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "Career Library_Updated_0508.xlsx"
OUT_DIR = ROOT / "prisma" / "seed-data" / "career-library"
OUT_DIR.mkdir(parents=True, exist_ok=True)


def s(v):
    """Normalize a cell value to a trimmed string, or None."""
    if v is None:
        return None
    v = str(v).strip()
    return v if v else None


# Known source-data spelling/naming variants that would otherwise silently break the
# join between tables (verified against the actual workbook content — see
# docs/db-design.md "Cross-table mapping"). CL is treated as the authoritative
# vocabulary; the other tabs' values are normalized to match it.
INDUSTRY_ALIASES = {"Defense": "Defence"}  # UG Institutions_IND -> CL spelling
EXAM_ALIASES = {"CUET": "CUET UG"}  # CL's extracted list -> UG Entrance_IND's exam name


def split_list(v, sep=","):
    if not v:
        return []
    return [p.strip() for p in str(v).split(sep) if p.strip()]


def parse_india_salary(text):
    """'10–25 LPA' -> (10.0, 25.0); non-numeric bounds (e.g. 'Limitless') -> None."""
    if not text:
        return None, None
    t = text.replace("₹", "").replace("LPA", "").strip()
    parts = re.split(r"[–‒\-]", t, maxsplit=1)
    if len(parts) != 2:
        return None, None

    def num(s_):
        s_ = s_.strip()
        return float(s_) if re.match(r"^\d+(\.\d+)?$", s_) else None

    return num(parts[0]), num(parts[1])


def parse_global_salary(text):
    """'$70k–$160k' -> (70000.0, 160000.0); '$0–Millions' -> (0.0, None)."""
    if not text:
        return None, None
    t = text.replace("$", "").strip()
    parts = re.split(r"[–‒\-]", t, maxsplit=1)
    if len(parts) != 2:
        return None, None

    def num(s_):
        s_ = s_.strip().rstrip("+")
        m = re.match(r"^(\d+(?:\.\d+)?)k$", s_, re.I)
        if m:
            return float(m.group(1)) * 1_000
        m = re.match(r"^(\d+(?:\.\d+)?)m$", s_, re.I)
        if m:
            return float(m.group(1)) * 1_000_000
        if re.match(r"^\d+(\.\d+)?$", s_):
            return float(s_)
        return None

    return num(parts[0]), num(parts[1])


AI_GRADE_MAP = {"low": "LOW", "medium": "MEDIUM", "high": "HIGH", "very high": "VERY_HIGH"}


def export_career_library(wb):
    ws = wb["CL"]
    out = []
    skipped = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        cluster, industry, domain, job_role = s(r[0]), s(r[1]), s(r[2]), s(r[3])
        ai_grade_raw = s(r[4])
        qual_10_12 = s(r[10])
        if not (cluster and industry and domain and job_role and ai_grade_raw and qual_10_12):
            skipped += 1
            continue

        india_min, india_max = parse_india_salary(r[8])
        global_min, global_max = parse_global_salary(r[9])

        out.append(
            {
                "cluster": cluster,
                "industry": industry,
                "domain": domain,
                "jobRole": job_role,
                "aiResilienceGrade": AI_GRADE_MAP.get(ai_grade_raw.lower(), "MEDIUM"),
                "aiResilienceComment": s(r[5]) or "",
                "oneLineDescription": s(r[6]) or "",
                "topCompanies": split_list(r[7]),
                "salaryIndiaRangeText": s(r[8]),
                "salaryIndiaMinLPA": india_min,
                "salaryIndiaMaxLPA": india_max,
                "salaryGlobalRangeText": s(r[9]),
                "salaryGlobalMinUSD": global_min,
                "salaryGlobalMaxUSD": global_max,
                "qualification10th12th": qual_10_12,
                "qualificationGraduation": s(r[11]),
                "entranceExamsUGDescription": s(r[12]),
                "entranceExams": [EXAM_ALIASES.get(t, t) for t in split_list(r[13])],
                "qualificationPG": s(r[14]),
                "entranceExamsPG": split_list(r[15]),
                "certificationsStudent": split_list(r[16], ";"),
                "certificationsUG": split_list(r[17], ";"),
                "topCourses": split_list(r[18]),
            }
        )
    print(f"CL: {len(out)} rows exported, {skipped} skipped (missing required field)")
    return out


def export_ug_institutions(wb):
    ws = wb["UG Institutions_IND"]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        industry, name = s(r[0]), s(r[2])
        if not (industry and name):
            continue
        industry = INDUSTRY_ALIASES.get(industry, industry)
        out.append(
            {
                "industry": industry,
                "shortName": s(r[1]),
                "name": name,
                "city": s(r[3]),
                "state": s(r[4]),
                "type": s(r[5]),
                "category": s(r[6]),
                "programmesOffered": s(r[7]),
                "programmesOfferedAfterClass12": s(r[8]),
                "keyProgrammesOffered": s(r[9]),
                "primaryEntranceExams": s(r[10]),
                "nirfRanking": s(r[11]),
                "otherRankings": s(r[12]),
                "approxAnnualFee": s(r[13]),
                "approxPlacementCtc": s(r[14]),
                "website": s(r[15]),
            }
        )
    print(f"UG Institutions_IND: {len(out)} rows exported")
    return out


def export_ug_inst_uty(wb):
    ws = wb["UG Inst+Uty_IND"]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        name = s(r[1])
        if not name:
            continue
        out.append(
            {
                "shortName": s(r[0]),
                "name": name,
                "city": s(r[2]),
                "state": s(r[3]),
                "type": s(r[4]),
                "category": s(r[5]),
                "keyProgrammesOffered": s(r[6]),
                "primaryEntranceExams": s(r[7]),
                "nirfRanking": s(r[8]),
                "otherRankings": s(r[9]),
                "approxAnnualFee": s(r[10]),
                "approxPlacementCtc": s(r[11]),
                "website": s(r[12]),
            }
        )
    print(f"UG Inst+Uty_IND: {len(out)} rows exported")
    return out


def export_ug_entrance(wb):
    ws = wb["UG Entrance_IND"]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        exam_name = s(r[0])
        if not exam_name:
            continue
        out.append(
            {
                "examName": exam_name,
                "fullForm": s(r[1]),
                "conductingBody": s(r[2]),
                "level": s(r[3]),
                "applicableFor": s(r[4]),
                "subjectRequirements12th": s(r[5]),
                "applicationWindow": s(r[6]),
                "examMonth": s(r[7]),
                "resultMonth": s(r[8]),
                "examMode": s(r[9]),
                "frequency": s(r[10]),
                "approxAttemptsAllowed": s(r[11]),
                "officialWebsite": s(r[12]),
            }
        )
    print(f"UG Entrance_IND: {len(out)} rows exported")
    return out


def export_ug_courses(wb):
    ws = wb["UG Courses_IND"]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        course_name, cluster = s(r[0]), s(r[4])
        if not (course_name and cluster):
            continue
        out.append(
            {
                "courseName": course_name,
                "fullForm": s(r[1]),
                "level": s(r[2]),
                "durationYears": s(r[3]) if r[3] is None or isinstance(r[3], str) else str(r[3]),
                "careerCluster": cluster,
                "stream12thRequirements": s(r[5]),
                "minimumEligibility": s(r[6]),
                "entranceExamsPrimary": s(r[7]),
                "entranceExamsAlternate": s(r[8]),
                "topSpecialisations": s(r[9]),
                "topGovtColleges": s(r[10]),
                "topPrivateColleges": s(r[11]),
                "approxAnnualFeeRange": s(r[12]),
                "furtherStudyOptions": s(r[13]),
            }
        )
    print(f"UG Courses_IND: {len(out)} rows exported")
    return out


def export_pg_institutions(wb):
    ws = wb["PG Institutions_IND"]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        institution = s(r[1])
        if not institution:
            continue
        out.append(
            {
                "industry": s(r[0]),
                "institution": institution,
                "state": s(r[2]),
                "city": s(r[3]),
                "programTypes": s(r[4]),
                "website": s(r[5]),
            }
        )
    print(f"PG Institutions_IND: {len(out)} rows exported")
    return out


def export_pg_entrance(wb):
    ws = wb["PG Entrance_IND"]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        exam_name = s(r[0])
        if not exam_name:
            continue
        out.append({"examName": exam_name, "coursesForExam": s(r[1]), "officialWebsite": s(r[2])})
    print(f"PG Entrance_IND: {len(out)} rows exported")
    return out


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)

    exports = {
        "career-library.json": export_career_library(wb),
        "ug-institutions.json": export_ug_institutions(wb),
        "ug-institutions-universities.json": export_ug_inst_uty(wb),
        "ug-entrance-exams.json": export_ug_entrance(wb),
        "ug-courses.json": export_ug_courses(wb),
        "pg-institutions.json": export_pg_institutions(wb),
        "pg-entrance-exams.json": export_pg_entrance(wb),
    }

    for filename, data in exports.items():
        path = OUT_DIR / filename
        path.write_text(json.dumps(data, indent=2, ensure_ascii=False))
        print(f"Wrote {path} ({len(data)} rows)")


if __name__ == "__main__":
    main()
